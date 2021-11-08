from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from openpyxl import cell
from datetime import date
import os

class ExcelController():
    def __init__(self, dataSource):
        self.dataSource = dataSource
        dateOfToday = date.today().strftime("%Y%m%d")
        self.fileName = "dailyCheck_"+ dateOfToday + ".xlsx"
        self.rowPosition = 1
        self.indent = 1
        if os.path.isfile(self.fileName):
            self.__makeExcel(self.fileName)
        else:
            self.__makeExcel()

    def __makeExcel(self, fileName = None):
        # 워크북을 연다.
        if fileName is not None:
            self.fileName = fileName
            self.workBook = load_workbook(fileName)
        else:
            self.workBook = Workbook()
        # 시트 생성
        title = self.dataSource["DATABASE_INFO"]["DBUNQNAME"]+"_"+self.dataSource["INSTANCE_INFO"][1]["HOSTNAME"]
        self.workSheet = self.workBook.create_sheet(self.dataSource["DATABASE_INFO"]["DBUNQNAME"])
        self.workBook.active
        self.workSheet.title = self.dataSource["DATABASE_INFO"]["DBUNQNAME"]
        # self.workSheet = self.workBook[self.dataSource["DATABASE_INFO"]["DBUNQNAME"]]

        # 각 셀들의 너비를 지정한다.
        self.workSheet.column_dimensions["A"].width = 1.50
        self.workSheet.column_dimensions["B"].width = 4.25
        self.workSheet.column_dimensions["C"].width = 16.55
        self.workSheet.column_dimensions["D"].width = 16.55
        self.workSheet.column_dimensions["E"].width = 16.55
        self.workSheet.column_dimensions["F"].width = 16.55
        self.workSheet.column_dimensions["G"].width = 16.55
        self.workSheet.column_dimensions["H"].width = 16.55
        self.workSheet.column_dimensions["I"].width = 16.55
        self.workSheet.column_dimensions["J"].width = 16.55
        self.workSheet.column_dimensions["K"].width = 16.55
        self.workSheet.column_dimensions["L"].width = 16.55

    def closeExcel(self):
        self.workBook.__delitem__("Sheet") if "Sheet" in self.workBook else None
        self.workBook.save(self.fileName)
        self.workBook.close()

    def writeTitle(self):
        title = self.dataSource["DATABASE_INFO"]["DBUNQNAME"]
        self.rowPosition = self.__writeData__(title, self.rowPosition, 'MAINTITLE')

    def writeDbInfo(self):
        # 수정해야함.
        from Controllers.excel_header_define import DBINFO_COMMON_HEADER as HEADER
        data = dict()
        title = 'DATABASE INFO'
        data = self.dataSource["DATABASE_INFO"]

        self.rowPosition = self.__writeData__(title, self.rowPosition, 'TITLE')
        self.rowPosition = self.__writeData__(HEADER,self.rowPosition, 'HEADER')
        self.rowPosition = self.__writeData__(data,  self.rowPosition, 'DATASINGLE')

    def writeInstanceInfo(self):
        from Controllers.excel_header_define import DBINFO_INSTANCE_HEADER as HEADER
        data = self.dataSource["INSTANCE_INFO"]
        title = 'INSTANCE INFO'

        self.rowPosition = self.__writeData__(title,self.rowPosition, 'TITLE')
        self.rowPosition = self.__writeData__(HEADER, self.rowPosition, 'HEADER')
        self.rowPosition = self.__writeData__(data, self.rowPosition,'DATASET')

    def writeMemoryDiag(self):
        from Controllers.excel_header_define import MEMORY_HEADER as HEADER
        data = dict()
        data = self.dataSource["MEMORYSTAT"]
        title = 'SGA/PGA INFO'

        self.rowPosition = self.__writeData__(title,self.rowPosition, 'TITLE')
        self.rowPosition = self.__writeData__(HEADER,self.rowPosition, 'HEADER')
        self.rowPosition = self.__writeData__(data,self.rowPosition, 'DATASET')

    def writeSgaOperDiag(self):
        from Controllers.excel_header_define import SGA_OPERATION_HEADER as HEADER
        data = dict()
        title = 'SGA OPERATION DIAG'
        for index in self.dataSource["SGAOPER"]:
            data = self.dataSource["SGAOPER"]

        self.rowPosition = self.__writeData__(title,self.rowPosition, 'TITLE')
        self.rowPosition = self.__writeData__(HEADER,self.rowPosition, 'HEADER')
        self.rowPosition = self.__writeData__(data, self.rowPosition, 'DATASET WITH SUBSET LIST')

    def writeDataGuardDiag(self):
        from Controllers.excel_header_define import DATA_GUARD_HEADER as HEADER
        title = 'DATAGUARD DIAG'
        data = dict()
        data = self.dataSource["DATAGUARD"]
        self.rowPosition = self.__writeData__(title,self.rowPosition, 'TITLE')
        self.rowPosition = self.__writeData__(HEADER,self.rowPosition, 'HEADER')
        self.rowPosition = self.__writeData__(data, self.rowPosition, 'DATASET WITH SUBSET')

    def writeAsmDiag(self):
        from Controllers.excel_header_define import ASM_HEADER as HEADER
        title = "ASM DIAG"
        data = dict()
        data = self.dataSource["ASMSTAT"]

        self.rowPosition = self.__writeData__(title,self.rowPosition, 'TITLE')
        self.rowPosition = self.__writeData__(HEADER,self.rowPosition, 'HEADER')
        self.rowPosition = self.__writeData__(data, self.rowPosition, 'DATASET')

    def writeTablespaceDiag(self):
        from Controllers.excel_header_define import TABLESPACE_HEADER as HEADER
        title = "TABLESPACE DIAG"
        data = dict()
        data = self.dataSource["TABLESPACE"]
        self.rowPosition = self.__writeData__(title, self.rowPosition, 'TITLE')
        self.rowPosition = self.__writeData__(HEADER, self.rowPosition, 'HEADER')
        self.rowPosition = self.__writeData__(data, self.rowPosition, 'DATASET')

    def writeAlertDiag(self):
        from Controllers.excel_header_define import ALERT_HEADER as HEADER
        title = "ALERTLOG DIAG"
        data = dict()
        data = self.dataSource["ALERT"]

        self.rowPosition = self.__writeData__(title, self.rowPosition,'TITLE')
        self.rowPosition = self.__writeData__(HEADER, self.rowPosition, 'HEADER')
        self.rowPosition = self.__writeData__(data, self.rowPosition, 'DATASET WITH LIST SUBSET')


    def __writeData__(self, data, position, dataType):
        # dataType : MAINTITLE, TITLE, HEADER, DATA
        # 기존에서 2줄 내려가서 시작
        if dataType == 'MAINTITLE':
            # MainTitle
            yPosition = 2
            xPosition = 1
            cell = self.workSheet.cell(row=yPosition, column=xPosition)
            cell.value = data
            cell.font = Font(name='Calibri', size=18, bold=True)
        elif dataType == 'TITLE':
            # Title
            yPosition = position + 1
            xPosition = 2
            cell = self.workSheet.cell(row=yPosition, column=xPosition)
            cell.value = data
            cell.font = Font(name='Calibri', size=13, bold=True)
        elif dataType == 'HEADER':
            yPosition = position + 1
            xPosition = 3
            for headerSequence in range(0,len(data)):
                cell = self.workSheet.cell(row=yPosition, column=xPosition+headerSequence)
                cell.value=data[headerSequence]
                cell.font = Font(name='Calibri', size=11, bold=True)
        elif dataType == 'DATASINGLE':
            yPosition = position + 1
            xPosition = 3
            for dataKey in data.keys():
                cell = self.workSheet.cell(row=yPosition, column=xPosition)
                cell.value=data[dataKey]
                cell.font=Font(name='Calibri',size=10)
                xPosition = xPosition + 1
            yPosition = yPosition + 1
        elif dataType == 'DATASET':
            yPosition = position + 1
            xPosition = 3
            for firstKey in data.keys():
                for secondKey in data[firstKey]:
                    cell = self.workSheet.cell(row=yPosition, column=xPosition)
                    cell.value=data[firstKey][secondKey]
                    cell.font = Font(name = 'Calibri', size = 10)
                    xPosition = xPosition + 1
                yPosition = yPosition + 1
                xPosition = 3
        elif dataType == 'DATASET WITH SUBSET LIST':
            yPosition = position + 1
            xPosition = 3
            for dataKey in data.keys():
                cell = self.workSheet.cell(row=yPosition, column=xPosition)
                cell.value = dataKey
                cell.font = Font(name = 'Calibri', size = 10)
                xPosition = xPosition + 1
                for attKey in data[dataKey].keys():
                    cell = self.workSheet.cell(row=yPosition, column=xPosition)
                    cell.value = attKey
                    cell.font = Font(name = 'Calibri', size = 10)
                    xPosition = xPosition + 1
                    for index in range(0,len(data[dataKey][attKey])):
                        for keyValue in data[dataKey][attKey][index].keys():
                            cell = self.workSheet.cell(row=yPosition, column=xPosition)
                            cell.value = data[dataKey][attKey][index][keyValue]
                            # print(dataKey, attKey, cell.value)
                            cell.font = Font(name = 'Calibri', size = 10)
                            xPosition = xPosition + 1
                        yPosition = yPosition + 1
                        xPosition = 5
                    xPosition = 4
                xPosition = 3
        elif dataType == 'DATASET WITH SUBSET':
            yPosition = position + 1
            xPosition = 3
            for firstKey in data.keys():
                for secondKey in data[firstKey].keys():
                    for thirdKey in data[firstKey][secondKey].keys():
                        cell = self.workSheet.cell(row=yPosition, column=xPosition)
                        cell.value = data[firstKey][secondKey][thirdKey]
                        cell.font = Font(name = 'Calibri', size = 10)
                        xPosition = xPosition + 1
                    yPosition = yPosition + 1
                    xPosition = 3
        elif dataType == "DATASET WITH LIST SUBSET":
            yPosition = position + 1
            for firstKey in data.keys():   # 0,1 -- Instance Number
                xPosition = 3
                cell = self.workSheet.cell(row=yPosition, column=xPosition)
                cell.value = firstKey
                cell.font = Font(name = 'Calibri', size = 10)
                xPosition = xPosition + 1
                for index in data[firstKey].keys(): # 로그 순서 키 '0','1'...
                    for secondKey in data[firstKey][index].keys(): 
                        cell = self.workSheet.cell(row=yPosition, column=xPosition)
                        cell.value = data[firstKey][index][secondKey]
                        cell.font = Font(name = 'Calibri', size = 10)
                        xPosition = xPosition + 1
                        # cell = self.workSheet.cell(row=yPosition, column=xPosition)
                        # cell.value = data[firstKey][index][secondKey]
                        # cell.font = Font(name = 'Calibri', size = 10)
                        # yPosition = yPosition + 1
                        # xPosition = 3
                    yPosition = yPosition + 1
                    xPosition = 4
        return yPosition
