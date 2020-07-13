from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from datetime import date
import os.path
class GenExcel:
    def __init__(self, sourceData):
        # get Current Year/Month/day
        getToday = date.today()
        dateOfToday = getToday.strftime("%Y%m%d")
        self.fileName = "dailycheck_" + dateOfToday + ".xlsx"
        self.openType = "Unknown"
        if os.path.isfile(self.fileName):
            self.workBook = load_workbook(self.fileName)
            self.openType = "Old"
        else:
            self.workBook = Workbook()
            self.openType = "New"

        # self.workSheet = self.workBook.active
        z = sourceData
        self.__dataCollect(z)
        self.__setNewSheet()
        self.__writeDBName()
        self.__memoryStatic()
        self.__sgaOperStatic()
        self.__dataguardStatic()
        self.__tablespaceStatic()
        self.__asmStatic()
        self.__backupStatic()
        self.__alertStatic()
        self.__closeExcel()

    def __dataCollect(self, dataSet):
        self.dbName = dataSet['DBNAME']
        self.memoryData = dataSet['MEMORY']
        self.tablespaceData = dataSet['TBSDATA']
        self.backupData = dataSet['BKDATA']
        self.alertData = dataSet['ALERT']
        self.asmData = dataSet['ASM']
        self.sgaOPData = dataSet['SGAOP']
        self.dgData = dataSet['DG']

    def __setNewSheet(self):
        # 시트 새로 만들기
        self.workSheet = self.workBook.create_sheet(self.dbName)
        self.workBook.active
        self.workSheet.title = self.dbName

        self.workSheet.column_dimensions["B"].width = 4.57
        self.workSheet.column_dimensions["C"].width = 19.15
        self.workSheet.column_dimensions["D"].width = 19.15
        self.workSheet.column_dimensions["E"].width = 19.15
        self.workSheet.column_dimensions["F"].width = 19.15
        self.workSheet.column_dimensions["G"].width = 19.15
        self.workSheet.column_dimensions["H"].width = 19.15
        self.workSheet.column_dimensions["I"].width = 19.15
        self.workSheet.column_dimensions["J"].width = 19.15
        self.workSheet.column_dimensions["K"].width = 19.15

    def __writeDBName(self):
        self.rowPosition=3
        # DB 이름
        self.workSheet["B"+str(self.rowPosition)] = self.dbName
        setFont = self.workSheet["B"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=18, bold=True)


    def __memoryStatic(self):
        self.rowPosition = self.rowPosition + 2
        self.workSheet["B"+str(self.rowPosition)] = "MEMORY STATISTICS"
        setFont = self.workSheet["B"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=13, bold=True)

        keyValue = list(self.memoryData.keys())

        for x in range(0,len(self.memoryData)):
            self.rowPosition = self.rowPosition + 1
            self.workSheet["C"+str(self.rowPosition)] = keyValue[x]
            self.workSheet["D"+str(self.rowPosition)] = self.memoryData[keyValue[x]]
            setFont = self.workSheet["C"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11, bold=True)
            setFont = self.workSheet["D"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11)

    def __sgaOperStatic(self):
        self.rowPosition = self.rowPosition + 2
        self.workSheet["B"+str(self.rowPosition)] = "SGA OPERATION STATISTICS"
        setFont = self.workSheet["B"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri', size=13, bold=True)

        self.rowPosition = self.rowPosition + 1
        self.workSheet["C"+str(self.rowPosition)] = "COMPONENT"
        self.workSheet["D"+str(self.rowPosition)] = "START TIME"
        self.workSheet["E"+str(self.rowPosition)] = "END TIME"
        self.workSheet["F"+str(self.rowPosition)] = "OPER TYPE"
        self.workSheet["G"+str(self.rowPosition)] = "OPER MODE"
        self.workSheet["H"+str(self.rowPosition)] = "INITIAL MB"
        self.workSheet["I"+str(self.rowPosition)] = "TARGET MB"
        self.workSheet["J"+str(self.rowPosition)] = "FINAL MB"
        self.workSheet["K"+str(self.rowPosition)] = "STATUS"

        setFont = self.workSheet["C"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["D"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["E"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["F"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["G"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["H"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["I"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["J"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["K"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)

        self.rowPosition = self.rowPosition + 1
        # SGA 변경내역은 키값이 0이 아니므로 키값이 0 이 있는 아이를 찾아 존재하면 SGA 변경내역이 없는것으로 판단.
        if 0 in self.sgaOPData:
            self.workSheet["C"+str(self.rowPosition)] = "No SGA Operations during period"
            setFont = self.workSheet["C"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11)
        else :
            for keyValue in self.sgaOPData.keys():
                self.workSheet["C"+str(self.rowPosition)] = keyValue
                setFont = self.workSheet["C"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                for y in range(0,len(self.sgaOPData[keyValue])): # 각 풀에 대하여 변경된 횟수만큼 반복한다.
                    self.workSheet["D"+str(self.rowPosition)] = self.sgaOPData[keyValue][y]["STARTDATE"]
                    self.workSheet["E"+str(self.rowPosition)] = self.sgaOPData[keyValue][y]["ENDDATE"]
                    self.workSheet["F"+str(self.rowPosition)] = self.sgaOPData[keyValue][y]["OPERATION"]
                    self.workSheet["G"+str(self.rowPosition)] = self.sgaOPData[keyValue][y]["OPERTYPE"]
                    self.workSheet["H"+str(self.rowPosition)] = self.sgaOPData[keyValue][y]["INITIAL"]
                    self.workSheet["I"+str(self.rowPosition)] = self.sgaOPData[keyValue][y]["TARGET"]
                    self.workSheet["J"+str(self.rowPosition)] = self.sgaOPData[keyValue][y]["FINAL"]
                    self.workSheet["K"+str(self.rowPosition)] = self.sgaOPData[keyValue][y]["STATUS"]

                    setFont = self.workSheet["D"+str(self.rowPosition)]
                    setFont.font = Font(name='Calibri',size=11)
                    setFont = self.workSheet["E"+str(self.rowPosition)]
                    setFont.font = Font(name='Calibri',size=11)
                    setFont = self.workSheet["F"+str(self.rowPosition)]
                    setFont.font = Font(name='Calibri',size=11)
                    setFont = self.workSheet["G"+str(self.rowPosition)]
                    setFont.font = Font(name='Calibri',size=11)
                    setFont = self.workSheet["H"+str(self.rowPosition)]
                    setFont.font = Font(name='Calibri',size=11)
                    setFont = self.workSheet["I"+str(self.rowPosition)]
                    setFont.font = Font(name='Calibri',size=11)
                    setFont = self.workSheet["J"+str(self.rowPosition)]
                    setFont.font = Font(name='Calibri',size=11)
                    setFont = self.workSheet["K"+str(self.rowPosition)]
                    setFont.font = Font(name='Calibri',size=11)

                    self.rowPosition = self.rowPosition + 1

    def __dataguardStatic(self):
        self.rowPosition = self.rowPosition + 2
        self.workSheet["B"+str(self.rowPosition)] = "DATAGUARD STATISTICS"
        setFont = self.workSheet["B"+str(self.rowPosition)]
        setFont.font = Font(name="Calibri",size=13, bold=True)

        if self.dgData[0] == "NoData":
            self.rowPosition = self.rowPosition + 1
            self.workSheet["C"+str(self.rowPosition)] = "No Dataguard Configuration"
        else:
            self.rowPosition = self.rowPosition + 1
            self.workSheet["C"+str(self.rowPosition)] = "DB UNIQ NAME"
            self.workSheet["D"+str(self.rowPosition)] = "DB ROLE"
            self.workSheet["E"+str(self.rowPosition)] = "CURR TIMESTAMP"
            self.workSheet["F"+str(self.rowPosition)] = "DIFFERENCE"

            setFont = self.workSheet["C"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11, bold=True)
            setFont = self.workSheet["D"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11, bold=True)
            setFont = self.workSheet["E"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11, bold=True)
            setFont = self.workSheet["F"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11, bold=True)

            for x in range(0,len(self.dgData)):
                self.rowPosition = self.rowPosition + 1
                self.workSheet["C"+str(self.rowPosition)] = self.dgData[x]["DBUNIQNAME"]
                self.workSheet["D"+str(self.rowPosition)] = self.dgData[x]["ROLE"]
                self.workSheet["E"+str(self.rowPosition)] = self.dgData[x]["SCN"]
                if self.dgData[x]["ROLE"] == 'PRIMARY DATABASE':
                    self.workSheet["F"+str(self.rowPosition)] = "-"
                else:
                    self.workSheet["F"+str(self.rowPosition)] = self.dgData[x]["DIFF"]

                setFont = self.workSheet["C"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["D"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["E"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["F"+str(self.rowPosition)]

    def __tablespaceStatic(self):
        self.rowPosition = self.rowPosition + 2
        self.workSheet["B"+str(self.rowPosition)] = "TABLESPACE STATISTICS"
        setFont = self.workSheet["B"+str(self.rowPosition)]
        setFont.font = Font(name="Calibri",size=13, bold=True)

        self.rowPosition = self.rowPosition + 1
        self.workSheet["C"+str(self.rowPosition)] = "TABLESPACE NAME"
        self.workSheet["D"+str(self.rowPosition)] = "TOTAL SIZE(GB)"
        self.workSheet["E"+str(self.rowPosition)] = "USED SIZE(GB)"
        self.workSheet["F"+str(self.rowPosition)] = "USED PCT(%)"
        self.workSheet["G"+str(self.rowPosition)] = "WARNING LEVEL"

        setFont = self.workSheet["C"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["D"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["E"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["F"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["G"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)

        for x in range(0,len(self.tablespaceData)):
            self.rowPosition = self.rowPosition + 1
            self.workSheet["C"+str(self.rowPosition)] = self.tablespaceData[x]["NAME"]
            self.workSheet["D"+str(self.rowPosition)] = self.tablespaceData[x]["TOTAL"]
            self.workSheet["E"+str(self.rowPosition)] = self.tablespaceData[x]["USED"]
            self.workSheet["F"+str(self.rowPosition)] = self.tablespaceData[x]["PCT"]
            self.workSheet["G"+str(self.rowPosition)] = self.tablespaceData[x]["LEVEL"]

            setFont = self.workSheet["C"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11)
            setFont = self.workSheet["D"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11)
            setFont = self.workSheet["E"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11)
            setFont = self.workSheet["F"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11)
            setFont = self.workSheet["G"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11)

    def __asmStatic(self):
        self.rowPosition = self.rowPosition + 2
        self.workSheet["B"+str(self.rowPosition)] = "ASM STATISTICS"
        setFont = self.workSheet["B"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri', size=13, bold=True)

        self.rowPosition = self.rowPosition + 1
        self.workSheet["C"+str(self.rowPosition)] = "DG NAME"
        self.workSheet["D"+str(self.rowPosition)] = "TOTAL_MB"
        self.workSheet["E"+str(self.rowPosition)] = "USED_MB"
        self.workSheet["F"+str(self.rowPosition)] = "USABLE_FILE_MB"
        self.workSheet["G"+str(self.rowPosition)] = "USED PCT"
        self.workSheet["H"+str(self.rowPosition)] = "WARNING LEVEL"

        setFont = self.workSheet["C"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["D"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["E"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["F"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["G"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["H"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)

        if self.asmData[0] == "NoASM":
            self.rowPosition = self.rowPosition + 1
            self.workSheet["C"+str(self.rowPosition)] = "This system doesn't use ASM Storage"
            setFont = self.workSheet["C"+str(self.rowPosition)]
            setFont.font = Font(name='Calibri',size=11)
        else:
            for x in range(0,len(self.asmData)):
                self.rowPosition = self.rowPosition + 1
                self.workSheet["C"+str(self.rowPosition)] = self.asmData[x]["NAME"]
                self.workSheet["D"+str(self.rowPosition)] = self.asmData[x]["TOTAL"]
                self.workSheet["E"+str(self.rowPosition)] = self.asmData[x]["USE"]
                self.workSheet["F"+str(self.rowPosition)] = self.asmData[x]["USABLE"]
                self.workSheet["G"+str(self.rowPosition)] = self.asmData[x]["PCT"]
                self.workSheet["H"+str(self.rowPosition)] = self.asmData[x]["LEVEL"]

                setFont = self.workSheet["C"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["D"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["E"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["F"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["G"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["H"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)

    def __backupStatic(self):
        self.rowPosition = self.rowPosition + 2
        self.workSheet["B"+str(self.rowPosition)] = "BACKUP STATISTICS"
        setFont = self.workSheet["B"+str(self.rowPosition)]
        setFont.font = Font(size=13, bold=True)

        self.rowPosition = self.rowPosition + 1
        self.workSheet["C"+str(self.rowPosition)] = "START DATE"
        self.workSheet["D"+str(self.rowPosition)] = "BACKUP SIZE(GB)"
        self.workSheet["E"+str(self.rowPosition)] = "BACKUP STATUS"
        setFont = self.workSheet["C"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["D"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)
        setFont = self.workSheet["E"+str(self.rowPosition)]
        setFont.font = Font(name='Calibri',size=11, bold=True)

        if self.backupData[0] == "NoBackup":
            self.rowPosition = self.rowPosition + 1
            self.workSheet["C"+str(self.rowPosition)] = "No RMAN Backup during period"
        else:
            for x in range(0,len(self.backupData)):
                self.rowPosition = self.rowPosition + 1
                self.workSheet["C"+str(self.rowPosition)] = self.backupData[x]["STARTDATE"]
                self.workSheet["D"+str(self.rowPosition)] = self.backupData[x]["SIZE"]
                self.workSheet["E"+str(self.rowPosition)] = self.backupData[x]["STATUS"]

                setFont = self.workSheet["C"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["D"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["E"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)

    def __alertStatic(self):
        self.rowPosition = self.rowPosition + 2
        self.workSheet["B"+str(self.rowPosition)] = "ALERTLOG STATISTICS"
        setFont = self.workSheet["B"+str(self.rowPosition)]
        setFont.font = Font(size=13, bold=True)

        self.rowPosition = self.rowPosition + 1
        self.workSheet["C"+str(self.rowPosition)] = "LOG DATE"
        self.workSheet["D"+str(self.rowPosition)] = "LOG MESSAGE"
        setFont = self.workSheet["C"+str(self.rowPosition)]
        setFont.font = Font(size=11, bold=True)
        setFont = self.workSheet["D"+str(self.rowPosition)]
        setFont.font = Font(size=11, bold=True)

        if self.alertData[0] == "NoData":
            self.rowPosition = self.rowPosition + 1
            self.workSheet["C"+str(self.rowPosition)] = "No Errors during that period"
        else:
            for x in range(0,len(self.alertData)):
                self.rowPosition = self.rowPosition + 1
                self.workSheet["C"+str(self.rowPosition)] = self.alertData[x]["LOGDATE"]
                self.workSheet["D"+str(self.rowPosition)] = self.alertData[x]["LOGMESSAGE"]

                setFont = self.workSheet["C"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)
                setFont = self.workSheet["D"+str(self.rowPosition)]
                setFont.font = Font(name='Calibri',size=11)

    def __closeExcel(self):
        # 엑셀 닫기
        if self.openType == "New":
             del self.workBook["Sheet"]
        self.workBook.save(self.fileName)
        self.workBook.close()
